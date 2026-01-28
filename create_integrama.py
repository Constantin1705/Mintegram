import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Integramă"

# Set column width and row height for square cells
for col in range(1, 30):
    ws.column_dimensions[get_column_letter(col)].width = 3
for row in range(1, 40):
    ws.row_dimensions[row].height = 20

# Define colors
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Define border
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Center alignment
center = Alignment(horizontal='center', vertical='center')

# Solutions from the second image (each row)
solutions = [
    "P-B-O-A-I-WATER MELON-PRAF-GAND-IREN-OG-IA-IZ-CUPIORI-V-GHICIT-A-LUNA-TEMA-VA-URMAS-T-SARAZI-",
    "TRANIC-VERDE-DECONT-UC-HE-ROTAIO-ACE-DARNICIL-SAU-UN-BILE-CAL-STIMA-EL-EROU-ISA-ABC-ESTIC-PEA-INEL-OMAR-",
    "A-A-SA-UI-ARAT-E-S-CAR-CABRAGE-AVMC-IU-A-PREA-A-A-TI-O-V-MANER-G-SWE-I-ACOPERIS-I-"
]

# Grid data - Based on the crossword image
# Row, Col, Type (Y=Yellow/Definition, W=White/Empty, B=Black, I=Image)
grid_data = [
    # Row 1
    (1, 1, 'I'), (1, 2, 'Y', 'ARDEI\nROȘU\nPEPENE\nVERDE'), (1, 3, 'Y', 'P'),
    (1, 4, 'Y', 'FASOLE'), (1, 5, 'Y'), (1, 6, 'Y', 'MAN'), (1, 7, 'Y'),
    (1, 8, 'Y', 'THAT\n(inf.)'), (1, 9, 'Y', 'A'), (1, 10, 'Y', 'BACK\n(pop.)'),
    (1, 11, 'Y'), (1, 12, 'Y', 'I'),
    
    # Row 2
    (2, 2, 'Y'), (2, 3, 'Y', 'W'), (2, 4, 'Y'), (2, 5, 'Y', 'WAKE\nUP'),
    (2, 6, 'Y'), (2, 7, 'Y', 'RIGA\nFRASER'), (2, 8, 'Y'), (2, 9, 'Y', 'VANATA'),
    (2, 10, 'Y'), (2, 11, 'Y', 'CEAPA'), (2, 12, 'Y'),
    
    # Row 3
    (3, 2, 'Y', 'DUST'), (3, 3, 'W'), (3, 12, 'Y'),
    
    # Row 4
    (4, 2, 'Y', 'TRAIN'), (4, 3, 'W'), (4, 7, 'B'), (4, 8, 'Y', 'IDEA\nTHOUGHT'),
    (4, 9, 'W'), (4, 12, 'Y'),
    
    # Row 5
    (5, 2, 'W'), (5, 3, 'W'), (5, 6, 'Y', 'SLOGAN'), (5, 7, 'Y', 'I-'),
    (5, 8, 'Y', 'EGO'), (5, 9, 'W'), (5, 11, 'Y', 'SUNT'), (5, 12, 'W'),
    
    # Row 6
    (6, 6, 'Y', 'MOROCOV'), (6, 11, 'Y', 'COUCH'), (6, 12, 'B'),
    
    # Row 7
    (7, 2, 'Y', 'SMACK,\nTOUCH'), (7, 3, 'W'), (7, 5, 'Y', 'OVEN,\nFURNACE'),
    (7, 6, 'Y', 'RADIO\nCASET'), (7, 7, 'W'), (7, 11, 'W'),
    
    # Row 8
    (8, 2, 'Y', 'SUGAR'), (8, 3, 'W'), (8, 6, 'W'), (8, 7, 'W'),
    (8, 8, 'Y', 'MOON'), (8, 9, 'W'), (8, 11, 'W'),
    
    # Row 9
    (9, 1, 'Y', 'DOVLECEL'), (9, 2, 'Y'), (9, 3, 'Y', 'GUESSED'),
    (9, 4, 'Y', 'SUBJECT'), (9, 5, 'Y', 'ONLYTWO\nMOMENTS'),
    (9, 6, 'Y', 'H'), (9, 8, 'W'), (9, 9, 'Y', 'FERTILE\n(fem.)'),
    (9, 10, 'W'), (9, 11, 'Y', 'GASTRA\nVETE'), (9, 12, 'Y', 'T'),
    
    # Continue with more rows...
]

# Populate grid
for row_idx in range(1, 30):
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = center
        
# Apply specific formatting based on image
# Yellow cells with definitions
yellow_cells = [
    (1, 2), (1, 3), (1, 4), (1, 5), (1, 6), (1, 7), (1, 8), (1, 9), (1, 10), (1, 11), (1, 12),
    (2, 2), (2, 3), (2, 4), (2, 5), (2, 6), (2, 7), (2, 8), (2, 9), (2, 10), (2, 11), (2, 12),
    # Add more yellow cells...
]

for row, col in yellow_cells:
    cell = ws.cell(row=row, column=col)
    cell.fill = yellow

# Black cells
black_cells = [
    (4, 7), (6, 12), # Add more black cells based on image
]

for row, col in black_cells:
    cell = ws.cell(row=row, column=col)
    cell.fill = black

# White cells (empty for letters)
# These are left as white by default

# Add text to definition cells
definitions = {
    (1, 2): "ARDEI\nROȘU\nPEPENE\nVERDE",
    (1, 3): "P",
    (1, 4): "FASOLE",
    (1, 6): "MAN",
    (1, 8): "THAT\n(inf.)",
    (1, 9): "A",
    (1, 10): "BACK\n(pop.)",
    (1, 12): "I",
    (2, 3): "W",
    (2, 5): "WAKE\nUP",
    (2, 7): "RIGA\nFRASER",
    (2, 9): "VANATA",
    (2, 11): "CEAPA",
    (3, 2): "DUST",
    (4, 2): "TRAIN",
    (4, 8): "IDEA\nTHOUGHT",
    (5, 6): "SLOGAN",
    (5, 7): "I-",
    (5, 8): "EGO",
    (5, 11): "SUNT",
    (6, 6): "MOROCOV",
    (6, 11): "COUCH",
    (7, 2): "SMACK,\nTOUCH",
    (7, 5): "OVEN,\nFURNACE",
    (7, 6): "RADIO\nCASET",
    (8, 2): "SUGAR",
    (8, 8): "MOON",
    (9, 1): "DOVLECEL",
    (9, 3): "GUESSED",
    (9, 4): "SUBJECT",
    (9, 5): "ONLY TWO\nMOMENTS",
    (9, 6): "H",
    (9, 9): "FERTILE\n(fem.)",
    (9, 11): "GASTRA\nVETE",
    (9, 12): "T",
}

for (row, col), text in definitions.items():
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = Font(size=8, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Save workbook
wb.save('E:/mintegram/Mintegram/integrama.xlsx')
print("Excel file created successfully!")
